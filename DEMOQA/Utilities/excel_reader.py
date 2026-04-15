import os
from openpyxl.reader.excel import load_workbook
from Utilities.paths import TEST_DATA_DIR

def fileread(workbook, sheet_name):
    data = []
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            # skip completely empty rows
            if any(cell is not None for cell in row):
                data.append(list(row))
    else:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook: {workbook}")

    return data

def workbookload(filename):
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    file_path = os.path.join(TEST_DATA_DIR, filename)
    
    # ✅ Load Workbook
    workbook = load_workbook(file_path)
    return workbook